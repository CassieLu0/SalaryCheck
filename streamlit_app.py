import re
from datetime import datetime
from io import BytesIO

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# Page
# =========================
st.set_page_config(page_title="DSP Salary Summary Tool", layout="wide")
st.title("DSP Salary Summary Tool")

# =========================
# Constants
# =========================
SHEET_DELIVERY = "派费明细"
SHEET_ROUTE = "参数_线路明细"
SHEET_OFFSET = "冲抵明细"
SHEET_CLAIM = "理赔明细"

COL_DRIVER_ID = "快递员ID"
COL_DRIVER = "快递员"
COL_ROUTE = "区域/线路"
COL_WEIGHT = "结算重量lb"
COL_TASK = "任务号"
COL_STOP = "STOP序号"

# Weight tiers (no "超重件" word)
TIERS = ["0-5lb", "5-20lb", "20lb以上"]
BUCKETS = [
    ("0-5lb", 0, 5),         # [0,5)
    ("5-20lb", 5, 20),       # [5,20)
    ("20lb以上", 20, None),   # [20, +inf)
]

# =========================
# Helpers
# =========================
def safe_filename(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = re.sub(r"\s+", "", s)
    return s or "工资单"

def parse_fleet_and_period_from_filename(name: str):
    """
    filename example:
      SEA-ZEP-派费明细-09_02_2026-15_02_2026-GOFO.xlsx

    date format supports:
      - MM_DD_YYYY
      - DD_MM_YYYY  (your sample)
    """
    if not name:
        return None, None

    m_fleet = re.search(r"^(.*?)-派费明细-", name)
    m_period = re.search(r"(\d{2}_\d{2}_\d{4})-(\d{2}_\d{2}_\d{4})", name)

    fleet = m_fleet.group(1).strip() if m_fleet else None
    if not m_period:
        return fleet, None

    p1, p2 = m_period.group(1), m_period.group(2)

    ok = False
    for fmt in ("%m_%d_%Y", "%d_%m_%Y"):
        try:
            datetime.strptime(p1, fmt)
            datetime.strptime(p2, fmt)
            ok = True
            break
        except Exception:
            pass

    period = f"{p1}-{p2}" if ok else None
    return fleet, period

def weight_bucket(w):
    try:
        w = float(w)
    except Exception:
        w = 0.0
    for name, lb_min, lb_max in BUCKETS:
        if lb_max is None:
            if w >= lb_min:
                return name
        else:
            if lb_min <= w < lb_max:
                return name
    return BUCKETS[0][0]

# =========================
# Workbook builder
# =========================
def build_workbook(
    df_delivery: pd.DataFrame,
    df_route_price: pd.DataFrame,
    df_offset: pd.DataFrame,
    df_claim_raw: pd.DataFrame | None,
    max_claim_types: int = 20,
):
    """
    ✅ 改成：司机只有一行（跨线路合并）
    - 线路列：展示该司机本期跑过的所有线路（逗号拼接）
    - 件数：跨线路相加
    - 金额：按“各线路单价”分别计算后再汇总（Python 计算写死数值，不再用 Excel INDEX 公式）
    - 调整（理赔/冲抵）：按司机汇总写一次
    - 工资：应付 = 送货合计金额 + 调整合计金额（符号跟随源表）
    - 未匹配理赔：以 UNMATCHED::<name> 作为占位 did 也进入司机汇总
    """

    # -------------------------
    # 0) 校验派费必备列
    # -------------------------
    must_cols = [COL_DRIVER_ID, COL_DRIVER, COL_ROUTE, COL_WEIGHT, COL_TASK, COL_STOP]
    miss = [c for c in must_cols if c not in df_delivery.columns]
    if miss:
        raise ValueError(f"派费明细缺少列: {miss}")

    # -------------------------
    # 1) 价格表 -> price_map (route,bucket,ticket) => price
    # -------------------------
    pr = df_route_price.copy()
    pr["线路"] = pr["线路"].astype(str).str.strip()
    pr["档位"] = pr["档位"].astype(str).str.strip()
    pr["首票单价"] = pd.to_numeric(pr["首票单价"], errors="coerce")
    pr["联单单价"] = pd.to_numeric(pr["联单单价"], errors="coerce")

    if pr[["首票单价", "联单单价"]].isna().any().any():
        raise ValueError("单价表存在空值/非数字，无法生成。")

    price_map = {}
    for row in pr.itertuples(index=False):
        route = str(getattr(row, "线路")).strip()
        bucket = str(getattr(row, "档位")).strip()
        p_first = float(getattr(row, "首票单价"))
        p_link = float(getattr(row, "联单单价"))
        price_map[(route, bucket, "首票")] = p_first
        price_map[(route, bucket, "联单")] = p_link

    # -------------------------
    # 2) 派费明细：route / bucket / ticket
    # -------------------------
    df = df_delivery.copy()
    df[COL_DRIVER_ID] = df[COL_DRIVER_ID].astype(str).str.strip()
    df[COL_DRIVER] = df[COL_DRIVER].astype(str).str.strip()
    df["route"] = df[COL_ROUTE].astype(str).fillna("").str.strip()

    df["_weight"] = pd.to_numeric(df[COL_WEIGHT], errors="coerce").fillna(0.0)
    df["_bucket"] = df["_weight"].apply(weight_bucket)

    gkey = [COL_DRIVER_ID, COL_TASK, COL_STOP]
    df["_rank_in_stop"] = df.groupby(gkey).cumcount() + 1
    df["_ticket"] = df["_rank_in_stop"].eq(1).map({True: "首票", False: "联单"})

    # 先按【司机×线路×bucket×ticket】统计件数
    cnt_route = (
        df.groupby([COL_DRIVER_ID, COL_DRIVER, "route", "_bucket", "_ticket"], dropna=False)
          .size().reset_index(name="件数")
    )

    # 计算每个组合的金额（按该线路单价）
    def calc_amt(row):
        r = str(row["route"]).strip()
        b = str(row["_bucket"]).strip()
        t = str(row["_ticket"]).strip()  # "首票"/"联单"
        n = int(row["件数"])
        p = price_map.get((r, b, t), None)
        if p is None:
            # 没找到价：直接报错更安全（防止漏算）
            raise ValueError(f"单价表缺少：线路={r} 档位={b} 票型={t}")
        return n * float(p)

    cnt_route["金额"] = cnt_route.apply(calc_amt, axis=1)

    # 再按【司机×bucket×ticket】跨线路汇总件数与金额
    cnt_driver = (
        cnt_route.groupby([COL_DRIVER_ID, COL_DRIVER, "_bucket", "_ticket"], dropna=False)
                 .agg(件数=("件数", "sum"), 金额=("金额", "sum"))
                 .reset_index()
    )

    # driver -> routes list（用于展示）
    driver_routes = (
        df.groupby([COL_DRIVER_ID])[ "route" ]
          .apply(lambda s: sorted({x.strip() for x in s.astype(str).tolist() if x and str(x).strip()}))
          .to_dict()
    )

    # driver -> most common name（兜底）
    did_to_name = (
        df_delivery[[COL_DRIVER_ID, COL_DRIVER]]
        .dropna()
        .assign(**{
            COL_DRIVER_ID: lambda x: x[COL_DRIVER_ID].astype(str).str.strip(),
            COL_DRIVER: lambda x: x[COL_DRIVER].astype(str).str.strip(),
        })
        .groupby(COL_DRIVER_ID)[COL_DRIVER]
        .agg(lambda s: s.value_counts().index[0])
        .to_dict()
    )

    # -------------------------
    # 3) 冲抵汇总（按司机ID；符号跟随源表）
    # -------------------------
    offset_cnt, offset_amt = {}, {}
    offset_dids = set()
    if df_offset is not None and not df_offset.empty:
        need_off = ["快递员ID", "费用合计_未税"]
        miss_off = [c for c in need_off if c not in df_offset.columns]
        if miss_off:
            raise ValueError(f"冲抵明细缺少列: {miss_off}")

        dfo = df_offset.copy()
        dfo["快递员ID"] = dfo["快递员ID"].astype(str).str.strip()
        dfo["费用合计_未税"] = pd.to_numeric(dfo["费用合计_未税"], errors="coerce").fillna(0.0)

        g_off = dfo.groupby("快递员ID", dropna=False).agg(
            cnt=("费用合计_未税", "size"),
            amt=("费用合计_未税", "sum"),
        )
        offset_cnt = g_off["cnt"].to_dict()
        offset_amt = g_off["amt"].to_dict()
        offset_dids = set(str(x).strip() for x in g_off.index.tolist())

    # -------------------------
    # 4) 理赔汇总（按司机ID；未匹配用 UNMATCHED::<name>）
    # -------------------------
    claim_cnt, claim_amt = {}, {}
    claim_dids = set()
    unmatched_claim = pd.DataFrame()

    name_to_id = (
        df_delivery[[COL_DRIVER, COL_DRIVER_ID]]
        .dropna()
        .assign(**{
            COL_DRIVER: lambda x: x[COL_DRIVER].astype(str).str.strip(),
            COL_DRIVER_ID: lambda x: x[COL_DRIVER_ID].astype(str).str.strip(),
        })
        .groupby(COL_DRIVER)[COL_DRIVER_ID]
        .agg(lambda s: s.value_counts().index[0])
        .to_dict()
    )

    def normalize_claim_type(x: str) -> str:
        x = str(x).strip()
        if x == "轨迹断更":
            return "断更"
        if x == "虚假签收":
            return "虚假签收"
        return x or "其他"

    if df_claim_raw is not None and not df_claim_raw.empty:
        need_cl = ["快递员", "理赔类型", "费用合计_未税"]
        miss_cl = [c for c in need_cl if c not in df_claim_raw.columns]
        if miss_cl:
            raise ValueError(f"理赔明细缺少列: {miss_cl}")

        dfc = df_claim_raw.copy()
        dfc["_name"] = dfc["快递员"].astype(str).str.strip()
        dfc["_dtype"] = dfc["理赔类型"].apply(normalize_claim_type)
        dfc["_did"] = dfc["_name"].map(name_to_id)
        dfc["费用合计_未税"] = pd.to_numeric(dfc["费用合计_未税"], errors="coerce").fillna(0.0)

        unmatched_claim = dfc[dfc["_did"].isna()].copy()

        # matched
        dfc_ok = dfc[dfc["_did"].notna()].copy()
        dfc_ok["_did"] = dfc_ok["_did"].astype(str).str.strip()

        g_cl = dfc_ok.groupby(["_did", "_dtype"], dropna=False).agg(
            cnt=("费用合计_未税", "size"),
            amt=("费用合计_未税", "sum"),
        )
        for (did, dtype), row in g_cl.iterrows():
            did_s = str(did).strip()
            dtype_s = str(dtype).strip()
            claim_cnt[(did_s, dtype_s)] = int(row["cnt"])
            claim_amt[(did_s, dtype_s)] = float(row["amt"])
        claim_dids |= set(dfc_ok["_did"].unique().tolist())

        # unmatched -> pseudo did
        if not unmatched_claim.empty:
            um = unmatched_claim.copy()
            um["_pseudo_did"] = "UNMATCHED::" + um["_name"].astype(str).str.strip()
            g_um = um.groupby(["_pseudo_did", "_dtype"], dropna=False).agg(
                cnt=("费用合计_未税", "size"),
                amt=("费用合计_未税", "sum"),
            )
            for (pseudo_did, dtype), row in g_um.iterrows():
                did_s = str(pseudo_did).strip()
                dtype_s = str(dtype).strip()
                claim_cnt[(did_s, dtype_s)] = claim_cnt.get((did_s, dtype_s), 0) + int(row["cnt"])
                claim_amt[(did_s, dtype_s)] = claim_amt.get((did_s, dtype_s), 0.0) + float(row["amt"])
            claim_dids |= set(g_um.index.get_level_values(0).astype(str).tolist())

    # 动态理赔类型（可限制数量）
    claim_types = sorted({dtype for (_, dtype) in claim_cnt.keys()})
    if len(claim_types) > max_claim_types:
        type_sum_abs = {}
        for (_did, dtype), amt in claim_amt.items():
            type_sum_abs[dtype] = type_sum_abs.get(dtype, 0.0) + abs(float(amt))
        top_types = [t for t, _ in sorted(type_sum_abs.items(), key=lambda x: x[1], reverse=True)[:max_claim_types]]

        for (did, dtype) in list(claim_cnt.keys()):
            if dtype not in top_types:
                claim_cnt[(did, "其他")] = claim_cnt.get((did, "其他"), 0) + claim_cnt[(did, dtype)]
                claim_amt[(did, "其他")] = claim_amt.get((did, "其他"), 0.0) + claim_amt[(did, dtype)]
                del claim_cnt[(did, dtype)]
                del claim_amt[(did, dtype)]

        claim_types = top_types
        if "其他" in {dtype for (_, dtype) in claim_cnt.keys()} and "其他" not in claim_types:
            claim_types.append("其他")

    # did_to_name 补充 unmatched 显示名
    for did in list(claim_dids):
        did_s = str(did)
        if did_s.startswith("UNMATCHED::"):
            nm = did_s.split("UNMATCHED::", 1)[1].strip() or "UNKNOWN"
            did_to_name[did_s] = f"{nm} (UNMATCHED)"

    # -------------------------
    # 5) 司机集合：派费司机 ∪ 冲抵司机 ∪ 理赔司机（含UNMATCHED）
    #     且司机只一行
    # -------------------------
    delivery_dids = set(df[COL_DRIVER_ID].unique().tolist())
    all_dids = sorted(set(delivery_dids) | offset_dids | claim_dids)

    # -------------------------
    # 6) Workbook + 参数_线路明细 + 明细sheet（便于对账）
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "司机汇总"

    ws_route = wb.create_sheet(SHEET_ROUTE)
    ws_route.append(["线路", "档位", "首票单价", "联单单价"])
    for cell in ws_route[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _, row in pr.iterrows():
        ws_route.append([row["线路"], row["档位"], float(row["首票单价"]), float(row["联单单价"])])

    def write_df_to_sheet(book: Workbook, name: str, dfx: pd.DataFrame):
        wsx = book.create_sheet(name)
        wsx.append(list(dfx.columns))
        for cell in wsx[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in dfx.itertuples(index=False, name=None):
            wsx.append(list(row))

    if df_offset is not None and not df_offset.empty:
        write_df_to_sheet(wb, SHEET_OFFSET, df_offset)
    if df_claim_raw is not None and not df_claim_raw.empty:
        write_df_to_sheet(wb, SHEET_CLAIM, df_claim_raw)
    if unmatched_claim is not None and not unmatched_claim.empty:
        write_df_to_sheet(wb, "理赔_未匹配", unmatched_claim)

    # -------------------------
    # 7) 样式 & 表头
    # -------------------------
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=16)
    fill_head = PatternFill("solid", fgColor="F2D6C9")
    fill_sub = PatternFill("solid", fgColor="F7E6D8")
    fill_blue = PatternFill("solid", fgColor="D9E1F2")

    # 送货 headers
    headers = []
    for b in [x[0] for x in BUCKETS]:
        headers.append((b, "首票"))
        headers.append((b, "联单"))

    # 调整 blocks：理赔类型 + 冲抵 + 合计
    ded_blocks = claim_types + ["冲抵", "合计"]

    last_col_guess = 5 + len(headers) * 2 + 2 + len(ded_blocks) * 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_guess)
    ws.cell(1, 1).value = "司机工资单汇总（司机维度合并线路）"
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).alignment = center

    ws.merge_cells("A3:A5"); ws["A3"].value = "顺号"
    ws.merge_cells("B3:B5"); ws["B3"].value = "快递员"
    ws.merge_cells("C3:D3"); ws["C3"].value = "应付工资"
    ws.merge_cells("C4:C5"); ws["C4"].value = "金额"
    ws.merge_cells("D4:D5"); ws["D4"].value = "件数"
    ws.merge_cells("E3:E5"); ws["E3"].value = "线路(合并展示)"

    start_col = 6
    end_delivery_col = start_col + len(headers) * 2 - 1
    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_delivery_col)
    ws.cell(3, start_col).value = "送货工资（跨线路合并）"

    col = start_col
    for (bucket, ticket) in headers:
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.cell(4, col).value = f"{bucket}-{ticket}"
        ws.cell(5, col).value = "件数"
        ws.cell(5, col + 1).value = "金额"
        col += 2

    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    ws.cell(4, col).value = "送货合计"
    ws.cell(5, col).value = "件数"
    ws.cell(5, col + 1).value = "金额"
    delivery_total_cnt_col = col
    delivery_total_amt_col = col + 1
    col += 2

    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + len(ded_blocks) * 2 - 1)
    ws.cell(3, col).value = "工资调整（符号跟随源表）"

    for i, name in enumerate(ded_blocks):
        ws.merge_cells(start_row=4, start_column=col + i * 2, end_row=4, end_column=col + i * 2 + 1)
        ws.cell(4, col + i * 2).value = name
        ws.cell(5, col + i * 2).value = "件数"
        ws.cell(5, col + i * 2 + 1).value = "金额"

    ded_start_col = col
    ded_total_cnt_col = col + (len(ded_blocks) - 1) * 2
    ded_total_amt_col = ded_total_cnt_col + 1

    for rr in range(3, 6):
        ws.row_dimensions[rr].height = 22
        for cc in range(1, ded_total_amt_col + 1):
            cell = ws.cell(rr, cc)
            cell.alignment = center
            cell.font = bold
            cell.border = border
            cell.fill = fill_head if rr == 3 else fill_sub

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 26

    # -------------------------
    # 8) 写数据行（司机一行）
    # -------------------------
    # 方便查询：driver,bucket,ticket -> (cnt,amt)
    lookup = {}
        for _, rr in cnt_driver.iterrows():
    did = str(rr[COL_DRIVER_ID]).strip()
    bucket = str(rr["_bucket"]).strip()
    ticket = str(rr["_ticket"]).strip()
    lookup[(did, bucket, ticket)] = (int(rr["件数"]), float(rr["金额"]))


    row_start = 6
    for idx, did in enumerate(all_dids, start=1):
        did = str(did).strip()
        r = row_start + idx - 1

        name = did_to_name.get(did, "UNKNOWN")
        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = f"{name} ({did})"

        rts = driver_routes.get(did, [])
        ws.cell(r, 5).value = ", ".join(rts) if rts else ""

        c = start_col
        delivery_cnt_cells = []
        delivery_amt_cells = []

        for bucket, ticket in headers:
            n, a = lookup.get((did, bucket, ticket), (0, 0.0))
            ws.cell(r, c).value = n
            ws.cell(r, c + 1).value = float(a)
            delivery_cnt_cells.append(f"{get_column_letter(c)}{r}")
            delivery_amt_cells.append(f"{get_column_letter(c+1)}{r}")
            c += 2

        ws.cell(r, delivery_total_cnt_col).value = f"=SUM({','.join(delivery_cnt_cells)})"
        ws.cell(r, delivery_total_amt_col).value = f"=SUM({','.join(delivery_amt_cells)})"

        # 调整块：理赔类型动态列 + 冲抵 + 合计
        colp = ded_start_col

        for t in claim_types:
            ws.cell(r, colp).value = claim_cnt.get((did, t), 0)
            ws.cell(r, colp + 1).value = float(claim_amt.get((did, t), 0.0))
            colp += 2

        ws.cell(r, colp).value = int(offset_cnt.get(did, 0))
        ws.cell(r, colp + 1).value = float(offset_amt.get(did, 0.0))
        colp += 2

        # 合计（除最后“合计”块本身外）
        cnt_terms, amt_terms = [], []
        for j in range(ded_start_col, ded_start_col + (len(ded_blocks) - 1) * 2, 2):
            cnt_terms.append(f"{get_column_letter(j)}{r}")
            amt_terms.append(f"{get_column_letter(j+1)}{r}")
        ws.cell(r, ded_total_cnt_col).value = f"=SUM({','.join(cnt_terms)})"
        ws.cell(r, ded_total_amt_col).value = f"=SUM({','.join(amt_terms)})"

        # ✅ 工资公式：数字是什么就怎么算（调整合计可正可负）
        ws.cell(r, 4).value = f"={get_column_letter(delivery_total_cnt_col)}{r}"
        ws.cell(r, 3).value = f"={get_column_letter(delivery_total_amt_col)}{r}+{get_column_letter(ded_total_amt_col)}{r}"

        # 样式
        for cc in range(1, ded_total_amt_col + 1):
            cell = ws.cell(r, cc)
            cell.alignment = center
            cell.border = border
            if cc == 3:
                cell.fill = fill_blue

    ws.freeze_panes = "A6"
    return wb
# =========================
# Session state
# =========================
st.session_state.setdefault("generated", False)
st.session_state.setdefault("output_bytes", None)
st.session_state.setdefault("output_name", None)
st.session_state.setdefault("last_uploaded_name", None)

# =========================
# Upload
# =========================
uploaded_file = st.file_uploader("Upload salary Excel (xlsx)", type=["xlsx"])
if uploaded_file is None:
    st.info("请先上传 xlsx 文件。")
    st.stop()

# New file => reset
if st.session_state["last_uploaded_name"] != uploaded_file.name:
    st.session_state["last_uploaded_name"] = uploaded_file.name
    st.session_state["generated"] = False
    st.session_state["output_bytes"] = None
    st.session_state["output_name"] = None

# Parse fleet & period (read-only)
fleet_name, period_str = parse_fleet_and_period_from_filename(uploaded_file.name)
if not fleet_name or not period_str:
    st.error(
        "文件名无法解析【车队名称/帐期范围】。\n"
        "请确保文件名格式：<车队>-派费明细-MM_DD_YYYY-MM_DD_YYYY-...xlsx（也支持DD_MM_YYYY）\n"
        f"当前文件名：{uploaded_file.name}"
    )
    st.stop()

st.caption(f"车队：{fleet_name} ｜ 帐期：{period_str}")

# =========================
# Only extract routes (light read)
# =========================
try:
    df_routes_only = pd.read_excel(uploaded_file, sheet_name=SHEET_DELIVERY, usecols=[COL_ROUTE])
    df_routes_only.columns = df_routes_only.columns.astype(str).str.strip()
except Exception as e:
    st.error(f"无法读取Excel或缺少列 {COL_ROUTE}：{e}")
    st.stop()

routes = sorted([r for r in df_routes_only[COL_ROUTE].astype(str).fillna("").str.strip().unique().tolist() if r])
if not routes:
    st.error("未从 派费明细 的“区域/线路”提取到任何线路，请检查源表。")
    st.stop()

st.caption(f"已识别线路数：{len(routes)}")

# =========================
# Price form (submit then run heavy)
# =========================
base_rows = [{"线路": r, "档位": t, "首票单价": 0.0, "联单单价": 0.0} for r in routes for t in TIERS]
df_price = pd.DataFrame(base_rows)

with st.form("price_form", clear_on_submit=False):
    st.subheader("线路单价（填完后点确认才生成账单）")
    edited_df = st.data_editor(
        df_price,
        use_container_width=True,
        hide_index=True,
        column_config={
            "线路": st.column_config.TextColumn("线路", disabled=True),
            "档位": st.column_config.TextColumn("档位", disabled=True),
            "首票单价": st.column_config.NumberColumn("首票单价", min_value=0.0, step=0.01, format="%.2f"),
            "联单单价": st.column_config.NumberColumn("联单单价", min_value=0.0, step=0.01, format="%.2f"),
        },
    )
    submit = st.form_submit_button("✅ 确认生成账单", type="primary", use_container_width=True)

# =========================
# Heavy work only after submit
# =========================
if submit:
    df_route_price = edited_df.copy()
    df_route_price["首票单价"] = pd.to_numeric(df_route_price["首票单价"], errors="coerce")
    df_route_price["联单单价"] = pd.to_numeric(df_route_price["联单单价"], errors="coerce")

    missing_price = df_route_price["首票单价"].isna() | df_route_price["联单单价"].isna()
    if missing_price.any():
        st.error("单价未填写完整或存在非数字，无法生成账单。")
        st.stop()

    # Read full workbook
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names
    except Exception as e:
        st.error(f"无法读取Excel：{e}")
        st.stop()

    if SHEET_DELIVERY not in sheet_names:
        st.error(f"缺少工作表：{SHEET_DELIVERY}；当前sheet：{sheet_names}")
        st.stop()

    df_delivery = pd.read_excel(uploaded_file, sheet_name=SHEET_DELIVERY)
    df_delivery.columns = df_delivery.columns.astype(str).str.strip()

    df_offset = pd.DataFrame()
    if SHEET_OFFSET in sheet_names:
        df_offset = pd.read_excel(uploaded_file, sheet_name=SHEET_OFFSET)
        df_offset.columns = df_offset.columns.astype(str).str.strip()

    df_claim_raw = None
    if SHEET_CLAIM in sheet_names:
        df_claim_raw = pd.read_excel(uploaded_file, sheet_name=SHEET_CLAIM)
        df_claim_raw.columns = df_claim_raw.columns.astype(str).str.strip()

    try:
        wb = build_workbook(
            df_delivery=df_delivery,
            df_route_price=df_route_price,
            df_offset=df_offset,
            df_claim_raw=df_claim_raw,
        )
    except Exception as e:
        st.error(f"生成失败：{e}")
        st.stop()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    out_name = f"{safe_filename(fleet_name)}-{period_str}-工资单.xlsx"
    st.session_state["generated"] = True
    st.session_state["output_bytes"] = buf.getvalue()
    st.session_state["output_name"] = out_name

    st.success("账单已生成，可以下载了。")

# =========================
# Download (green after generated)
# =========================
if st.session_state.get("generated") and st.session_state.get("output_bytes"):
    st.markdown(
        """
        <style>
        div[data-testid="stDownloadButton"] button {
            background-color: #16a34a !important;
            color: white !important;
            border: 1px solid #15803d !important;
        }
        div[data-testid="stDownloadButton"] button:hover {
            background-color: #15803d !important;
            border-color: #166534 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.download_button(
        label="⬇️ 下载工资单",
        data=st.session_state["output_bytes"],
        file_name=st.session_state["output_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
else:
    st.info("请先填写单价，然后点击「确认生成账单」。生成后会出现绿色下载按钮。")
